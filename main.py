from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, File, UploadFile, Form, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import json, os, re, secrets
import PyPDF2, docx2txt, io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from groq import Groq
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = FastAPI(title="Shiro AI HR", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

security = HTTPBasic()
HR_USERNAME = os.environ.get("HR_USERNAME", "hr_admin")
HR_PASSWORD = os.environ.get("HR_PASSWORD", "shiro@2026")
MAIL_USERNAME = os.environ.get("MAIL_USERNAME")
MAIL_PASSWORD = os.environ.get("MAIL_PASSWORD")
MAIL_FROM = os.environ.get("MAIL_FROM")
client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

JOBS = [
    {"id": 1, "title": "Business Development Associate", "location": "Gurgaon", "experience": "0-2 years", "description": "We are looking for energetic BDA candidates with strong sales acumen, excellent English communication skills, comfortable in target-driven environment, strong relationship building skills. Graduation mandatory. B2C sales or EdTech internship experience preferred."},
    {"id": 2, "title": "AI/ML Engineer", "location": "Chennai", "experience": "0-1 years", "description": "We are looking for fresher AI/ML engineers with Python and LLM experience. Skills required: Python, TensorFlow, PyTorch, scikit-learn, RAG pipelines, prompt engineering, LLM integration. Freshers welcome."},
    {"id": 3, "title": "Data Analyst", "location": "Chennai", "experience": "0-2 years", "description": "We are looking for data analysts with SQL, Python and Power BI skills. Must know data visualization, dashboard creation, DAX, Excel. Freshers welcome."},
]

def verify_hr(credentials: HTTPBasicCredentials = Depends(security)):
    ok_user = secrets.compare_digest(credentials.username.strip(), HR_USERNAME.strip())
    ok_pass = secrets.compare_digest(credentials.password.strip(), HR_PASSWORD.strip())
    if not (ok_user and ok_pass):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials", headers={"WWW-Authenticate": "Basic"})
    return credentials.username

def extract_text(file_bytes, filename):
    ext = filename.lower().split(".")[-1]
    try:
        if ext == "pdf":
            reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        elif ext in ("doc", "docx"):
            return docx2txt.process(io.BytesIO(file_bytes))
        else:
            return file_bytes.decode("utf-8", errors="ignore")
    except:
        return ""

def screen_candidate(jd_text, resume_text, filename, weights):
    prompt = f"""You are an expert HR recruiter. Evaluate this resume against the job description.

JOB DESCRIPTION:
{jd_text[:3000]}

RESUME:
{resume_text[:3000]}

WEIGHTS: Skills {weights.get('skills',40)}%, Experience {weights.get('experience',30)}%, Education {weights.get('education',15)}%, Certifications {weights.get('certifications',15)}%

Respond with ONLY this JSON (no markdown, no extra text):
{{"name":"candidate name","score":75,"breakdown":{{"skills":80,"experience":70,"education":75,"certifications":60}},"category":"Recommended","summary":"2 sentence summary."}}

Categories: Highly Recommended>=80, Recommended>=60, Consider>=40, Not Recommended<40"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3
    )
    raw = response.choices[0].message.content.strip()
    raw = re.sub(r"```json|```", "", raw).strip()
    result = json.loads(raw)
    result["filename"] = filename
    return result

def send_email(to_email, subject, body):
    try:
        msg = MIMEMultipart()
        msg["From"] = MAIL_FROM
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))
        with smtplib.SMTP_SSL("smtp.zoho.in", 465) as server:
            server.login(MAIL_USERNAME, MAIL_PASSWORD)
            server.sendmail(MAIL_FROM, to_email, msg.as_string())
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

# ── CANDIDATE PORTAL ENDPOINTS ────────────────────────────────

@app.get("/api/jobs")
async def get_jobs():
    return {"jobs": JOBS}

@app.post("/api/candidate/apply")
async def candidate_apply(
    candidate_name: str = Form(...),
    candidate_email: str = Form(...),
    job_id: int = Form(...),
    resume_file: UploadFile = File(...)
):
    job = next((j for j in JOBS if j["id"] == job_id), None)
    if not job:
        return JSONResponse(status_code=404, content={"error": "Job not found"})

    resume_bytes = await resume_file.read()
    resume_text = extract_text(resume_bytes, resume_file.filename)

    if not resume_text.strip():
        return JSONResponse(status_code=400, content={"error": "Could not extract text from resume"})

    weights = {"skills": 40, "experience": 30, "education": 15, "certifications": 15}
    try:
        result = screen_candidate(job["description"], resume_text, resume_file.filename, weights)
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

    # Email to HR
    hr_body = f"""
New Candidate Application Received!

Candidate: {candidate_name}
Email: {candidate_email}
Applied for: {job['title']}
Location: {job['location']}

AI Assessment:
Score: {result.get('score')}%
Category: {result.get('category')}
Summary: {result.get('summary')}

Score Breakdown:
- Skills: {result.get('breakdown', {}).get('skills')}%
- Experience: {result.get('breakdown', {}).get('experience')}%
- Education: {result.get('breakdown', {}).get('education')}%
- Certifications: {result.get('breakdown', {}).get('certifications')}%

Best regards,
Shiro AI HR System
PixelKliQ Technologies
    """
    send_email(MAIL_USERNAME, f"New Application: {candidate_name} for {job['title']}", hr_body)

    # Confirmation email to candidate
    candidate_body = f"""
Dear {candidate_name},

Thank you for applying for {job['title']} at PixelKliQ Technologies!

We have received your application and our AI screening system has processed your profile. Our HR team will review and get back to you shortly.

Best regards,
HR Team
PixelKliQ Technologies
    """
    send_email(candidate_email, f"Application Received - {job['title']} | PixelKliQ Technologies", candidate_body)

    return {
        "status": "success",
        "message": "Application submitted successfully!",
        "candidate": candidate_name,
        "job": job["title"],
        "score": result.get("score"),
        "category": result.get("category"),
        "summary": result.get("summary")
    }

# ── HR PORTAL ENDPOINTS ───────────────────────────────────────

@app.post("/api/screen")
async def screen_resumes(
    jd_file: UploadFile = File(...),
    resume_files: list[UploadFile] = File(...),
    weights: str = Form('{"skills":40,"experience":30,"education":15,"certifications":15}'),
    username: str = Depends(verify_hr)
):
    try:
        weights_dict = json.loads(weights)
    except:
        weights_dict = {"skills":40,"experience":30,"education":15,"certifications":15}

    jd_text = extract_text(await jd_file.read(), jd_file.filename)
    if not jd_text.strip():
        return JSONResponse(status_code=400, content={"error": "Could not extract JD text."})

    candidates = []
    for resume in resume_files:
        resume_text = extract_text(await resume.read(), resume.filename)
        if not resume_text.strip():
            candidates.append({"filename": resume.filename, "name": resume.filename, "score": 0, "category": "Not Recommended", "summary": "Could not extract text.", "breakdown": {"skills":0,"experience":0,"education":0,"certifications":0}})
            continue
        try:
            result = screen_candidate(jd_text, resume_text, resume.filename, weights_dict)
            candidates.append(result)
        except Exception as e:
            candidates.append({"filename": resume.filename, "name": resume.filename, "score": 0, "category": "Not Recommended", "summary": f"Error: {str(e)}", "breakdown": {"skills":0,"experience":0,"education":0,"certifications":0}})

    candidates.sort(key=lambda x: x.get("score", 0), reverse=True)
    return {"total": len(candidates), "candidates": candidates, "jd_filename": jd_file.filename}

@app.post("/api/send-email")
async def send_interview_email(data: dict, username: str = Depends(verify_hr)):
    to_email = data.get("email")
    candidate_name = data.get("name")
    job_title = data.get("job_title", "the position")
    interview_date = data.get("interview_date")
    interview_time = data.get("interview_time")

    if not to_email:
        return JSONResponse(status_code=400, content={"error": "Email address required"})

    body = f"""
Dear {candidate_name},

We are pleased to inform you that you have been shortlisted for {job_title} at PixelKliQ Technologies.

Interview Details:
Date: {interview_date}
Time: {interview_time}

Please confirm your availability by replying to this email.

Best regards,
HR Team
PixelKliQ Technologies
    """
    success = send_email(to_email, f"Interview Invitation - {job_title} | PixelKliQ Technologies", body)
    if success:
        return {"status": "sent", "message": f"Email sent to {to_email}"}
    else:
        return JSONResponse(status_code=500, content={"error": "Failed to send email"})

@app.post("/api/export-excel")
async def export_excel(data: dict, username: str = Depends(verify_hr)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Screening Results"
    header_fill = PatternFill(start_color="1E2130", end_color="1E2130", fill_type="solid")
    headers = ["Name","File","Score","Category","Skills","Experience","Education","Certifications","Summary"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="F5C842", bold=True)
        cell.alignment = Alignment(horizontal="center")
    colors_map = {"Highly Recommended":"00C896","Recommended":"4A9EFF","Consider":"F5A623","Not Recommended":"FF5C5C"}
    for row, c in enumerate(data.get("candidates", []), 2):
        color = colors_map.get(c.get("category",""), "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        vals = [c.get("name",""), c.get("filename",""), c.get("score",0), c.get("category",""), c.get("breakdown",{}).get("skills",0), c.get("breakdown",{}).get("experience",0), c.get("breakdown",{}).get("education",0), c.get("breakdown",{}).get("certifications",0), c.get("summary","")]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            if col == 4:
                cell.fill = fill
                cell.font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in col) + 4, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=ShiroHR_Results.xlsx"})

@app.get("/api/health")
def health():
    return {"status": "ok", "service": "Shiro AI HR"}

@app.post("/api/verify-login")
async def verify_login(username: str = Depends(verify_hr)):
    return {"status": "ok", "user": username}
